from pathlib import Path
import csv
import json
import re
import zipfile
from openpyxl import load_workbook

root = Path('thirdwave')
res = root / 'results'
schema_out = root / 'schema-out'
log_out = root / 'log-out'
converted = root / 'converted'
controls = root / 'controls-converted'

payloads = {
    'plain_eq': '=1+1',
    'plain_plus': '+11+11',
    'plain_minus': '-12+12',
    'plain_at': '@SUM(13,13)',
    'space': ' =2+2',
    'tab': '\t=3+3',
    'cr': '\r=4+4',
    'lf': '\n=5+5',
    'vt': '\x0b=6+6',
    'ff': '\x0c=7+7',
    'nbsp': '\u00a0=8+8',
    'bom': '\ufeff=9+9',
    'zwsp': '\u200b=10+10',
}

def files_under(base, suffix):
    return [p for p in base.rglob('*') if p.is_file() and p.suffix.lower() == suffix]

def inspect_csv(path):
    rows = list(csv.reader(path.open(newline='', errors='replace')))
    hits = []
    for ri, row in enumerate(rows, 1):
        for ci, cell in enumerate(row, 1):
            for label, payload in payloads.items():
                if cell == payload or payload in cell:
                    stripped = cell.lstrip('\ufeff \t\r\n\x0b\x0c\u00a0\u200b')
                    hits.append({
                        'file': str(path), 'row': ri, 'col': ci,
                        'label': label, 'payload': payload,
                        'cell_repr': repr(cell),
                        'exact': cell == payload,
                        'apostrophe_neutralized': stripped.startswith("'") and stripped[1:].lstrip().startswith(('=', '+', '-', '@')),
                        'formula_after_common_trim': stripped.startswith(('=', '+', '-', '@')),
                    })
    return {'path': str(path), 'rows': len(rows), 'hits': hits}

def inspect_workbooks(base):
    books = []
    for path in files_under(base, '.xlsx'):
        item = {'path': str(path), 'cells': [], 'error': None}
        try:
            wb = load_workbook(path, data_only=False, read_only=False)
            for ws in wb.worksheets:
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is None:
                            continue
                        value = str(cell.value)
                        formula = cell.data_type == 'f' or value.startswith('=')
                        matched = []
                        for label, payload in payloads.items():
                            canonical = payload.lstrip('\ufeff \t\r\n\x0b\x0c\u00a0\u200b')
                            if payload in value or canonical in value:
                                matched.append(label)
                        if formula or matched:
                            item['cells'].append({
                                'sheet': ws.title,
                                'coordinate': cell.coordinate,
                                'data_type': cell.data_type,
                                'value': value[:1000],
                                'formula': formula,
                                'matched_payloads': matched,
                            })
        except Exception as exc:
            item['error'] = repr(exc)
        books.append(item)
    return books

report = {
    'version': (res / 'version.txt').read_text(errors='replace').strip() if (res / 'version.txt').exists() else '',
    'schema_exit': (res / 'schema.exit').read_text().strip() if (res / 'schema.exit').exists() else None,
    'log_exit': (res / 'log.exit').read_text().strip() if (res / 'log.exit').exists() else None,
    'seed_log': (res / 'seed.log').read_text(errors='replace') if (res / 'seed.log').exists() else '',
    'schema_csv': [inspect_csv(p) for p in files_under(schema_out, '.csv')],
    'log_csv': [inspect_csv(p) for p in files_under(log_out, '.csv')],
    'trr_workbooks': inspect_workbooks(converted),
    'control_workbooks': inspect_workbooks(controls),
    'browser_dumps': {},
}
for p in files_under(res / 'browser-dumps', '.html'):
    text = p.read_text(errors='replace')
    report['browser_dumps'][str(p)] = {
        'path_exec': 'data-trr-log-path="executed"' in text.lower(),
        'attr_exec': 'data-trr-log-attr="executed"' in text.lower(),
        'marker_present': 'TRR_LOG_PATH' in text,
        'excerpt': text[:2000],
    }

control_formula_labels = set()
for book in report['control_workbooks']:
    for cell in book['cells']:
        if cell['formula']:
            control_formula_labels.update(cell['matched_payloads'])
trr_formula_labels = set()
trr_formula_cells = []
for book in report['trr_workbooks']:
    for cell in book['cells']:
        if cell['formula'] and cell['matched_payloads']:
            trr_formula_labels.update(cell['matched_payloads'])
            trr_formula_cells.append({'book': book['path'], **cell})
report['control_formula_labels'] = sorted(control_formula_labels)
report['trr_formula_labels'] = sorted(trr_formula_labels)
report['trr_formula_cells'] = trr_formula_cells
report['confirmed_csv_bypass_labels'] = sorted(control_formula_labels & trr_formula_labels - {'plain_eq', 'plain_plus', 'plain_minus', 'plain_at'})
report['plain_formula_labels_in_trr'] = sorted(trr_formula_labels & {'plain_eq', 'plain_plus', 'plain_minus', 'plain_at'})
report['log_path_html_execution'] = any(
    item['path_exec'] or item['attr_exec'] for item in report['browser_dumps'].values()
)

canaries = ['TRR_SECRET_CANARY_7A91FD', 'TRR_SECRET_CANARY_7A91FD_WRONG']
secret_hits = []
for p in [res / 'schema.stdout', res / 'schema.stderr', res / 'log.stdout', res / 'log.stderr', res / 'auth.stdout', res / 'auth.stderr']:
    if not p.exists():
        continue
    text = p.read_text(errors='replace')
    for canary in canaries:
        if canary in text:
            secret_hits.append({'file': str(p), 'canary': canary, 'excerpt': text[:3000]})
report['secret_hits'] = secret_hits

report['verdicts'] = {
    'csv_current_fix_bypass': 'confirmed' if report['confirmed_csv_bypass_labels'] else 'disproved',
    'csv_plain_known_class': 'present' if report['plain_formula_labels_in_trr'] else 'fixed',
    'log_path_html': 'confirmed' if report['log_path_html_execution'] else 'disproved',
    'credential_output': 'potential' if secret_hits else 'disproved',
}
(res / 'analysis.json').write_text(json.dumps(report, indent=2, ensure_ascii=False))

lines = [
    '# Agent 8 third-wave validation', '',
    f"Version: `{report['version']}`",
    f"Schema exit: `{report['schema_exit']}`",
    f"Log exit: `{report['log_exit']}`", '',
    '## Verdicts',
]
for key, value in report['verdicts'].items():
    lines.append(f'- **{key}: {value}**')
lines += [
    '',
    f"Control spreadsheet formula labels: `{report['control_formula_labels']}`",
    f"TRR spreadsheet formula labels: `{report['trr_formula_labels']}`",
    f"Confirmed current-fix bypass labels: `{report['confirmed_csv_bypass_labels']}`",
    f"Log-path browser execution: `{report['log_path_html_execution']}`",
    f"Credential hits: `{len(secret_hits)}`",
    '', '## Full evidence', '```json',
    json.dumps(report, indent=2, ensure_ascii=False),
    '```',
]
(res / 'analysis.md').write_text('\n'.join(lines))
