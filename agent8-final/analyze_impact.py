#!/usr/bin/env python3
from __future__ import annotations

import csv
import json
import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

root = Path("final-run")
meta = json.loads((root / "impact-meta.json").read_text(encoding="utf-8"))
result: dict[str, Any] = {
    "selected_label": meta.get("selected_label"),
    "available_labels": meta.get("available_labels", []),
    "skipped": bool(meta.get("skipped")),
    "run_ok": False,
    "formula_preserved": False,
    "hyperlink_formula_preserved": False,
    "automatic_callback": False,
    "control_formula_preserved": False,
    "direct_negative_is_formula": False,
    "normal_negative_is_formula": False,
    "raw_csv_hits": [],
    "trr_formula_cells": [],
    "control_formula_cells": [],
    "callbacks": [],
    "errors": [],
}
if result["skipped"]:
    (root / "csv-impact.json").write_text(json.dumps(result, indent=2), encoding="utf-8")
    raise SystemExit(0)

payloads = meta.get("payloads", {})
output = root / "impact-out"

for path in output.rglob("*.csv") if output.exists() else []:
    try:
        rows = list(csv.reader(path.open(newline="", errors="replace")))
    except Exception as exc:
        result["errors"].append(f"csv {path}: {exc!r}")
        continue
    for row_index, row in enumerate(rows, 1):
        for column_index, value in enumerate(row, 1):
            for label, payload in payloads.items():
                if payload in value:
                    result["raw_csv_hits"].append({
                        "file": str(path),
                        "row": row_index,
                        "column": column_index,
                        "label": label,
                        "cell_repr": repr(value),
                        "exact": value == payload,
                    })


def inspect_workbooks(base: Path, target: str) -> list[dict[str, Any]]:
    cells: list[dict[str, Any]] = []
    for path in base.rglob("*.xlsx") if base.exists() else []:
        try:
            workbook = load_workbook(path, data_only=False, read_only=False)
        except Exception as exc:
            result["errors"].append(f"xlsx {path}: {exc!r}")
            continue
        for sheet in workbook.worksheets:
            for row in sheet.iter_rows():
                for cell in row:
                    if cell.value is None:
                        continue
                    value = str(cell.value)
                    is_formula = cell.data_type == "f" or value.startswith("=")
                    matched = [
                        label for label, payload in payloads.items()
                        if payload in value
                        or payload.lstrip("\ufeff \t\r\n\x0b\x0c\u00a0\u200b") in value
                    ]
                    if is_formula or matched:
                        cells.append({
                            "target": target,
                            "file": str(path),
                            "sheet": sheet.title,
                            "coordinate": cell.coordinate,
                            "value": value[:2000],
                            "data_type": cell.data_type,
                            "is_formula": is_formula,
                            "matched": matched,
                        })
    return cells

result["trr_formula_cells"] = inspect_workbooks(root / "impact-converted", "trr")
result["control_formula_cells"] = inspect_workbooks(root / "control-converted", "control")

for cell in result["trr_formula_cells"]:
    if not cell["is_formula"]:
        continue
    matched = set(cell["matched"])
    if matched & {"arithmetic", "webservice", "hyperlink"}:
        result["formula_preserved"] = True
    if "hyperlink" in matched:
        result["hyperlink_formula_preserved"] = True
    if "direct_negative" in matched:
        result["direct_negative_is_formula"] = True
    if "normal_negative" in matched:
        result["normal_negative_is_formula"] = True
for cell in result["control_formula_cells"]:
    if cell["is_formula"] and set(cell["matched"]) & {"arithmetic", "webservice", "hyperlink"}:
        result["control_formula_preserved"] = True

callback_path = root / "results" / "impact-callbacks.log"
if callback_path.exists():
    result["callbacks"] = callback_path.read_text(errors="replace").splitlines()
result["automatic_callback"] = any("TRR_CSV_WEBSERVICE" in line for line in result["callbacks"])

run_exit = (root / "results" / "impact-run.exit")
seed_log = (root / "results" / "impact-seed.log")
result["run_exit"] = run_exit.read_text().strip() if run_exit.exists() else None
result["seed_ok"] = seed_log.exists() and "IMPACT_SEED_OK" in seed_log.read_text(errors="replace")
result["run_ok"] = (
    result["run_exit"] == "0"
    and result["seed_ok"]
    and bool(result["raw_csv_hits"])
    and result["control_formula_preserved"]
)

# The current direct-prefix CVE class is the negative control: it must not remain a formula.
result["negative_control_passed"] = not result["direct_negative_is_formula"] and not result["normal_negative_is_formula"]
result["independent_confirmation"] = bool(
    result["run_ok"]
    and result["formula_preserved"]
    and result["negative_control_passed"]
)

(root / "csv-impact.json").write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
